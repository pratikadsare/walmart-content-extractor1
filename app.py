from __future__ import annotations

import io
import json
import re
import time
from html import unescape
from dataclasses import dataclass, field
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
import requests
import streamlit as st
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

from auth import authenticate_user, get_display_name, normalize_email

APP_TITLE = "Walmart Content Extractor"
APP_SUBTITLE = (
    "Choose only the Walmart PDP attributes you need, then export a polished Excel file."
)
DEFAULT_ROWS = 10
MAX_ROWS = 1000
MAX_VISIBLE_ROWS = 20
SCROLL_AFTER_ROWS = 15
TABLE_ROW_HEIGHT = 40
NOT_FOUND_TEXT = "Unable to find on PDP"
DEFAULT_OUTPUT_FILENAME = "walmart_content_export"

INPUT_COLUMNS = ["SKU", "Walmart URL"]

REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36"
    ),
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,"
        "image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Cache-Control": "no-cache",
    "Pragma": "no-cache",
    "Referer": "https://www.walmart.com/",
    "Upgrade-Insecure-Requests": "1",
}

REQUEST_DELAY_SECONDS = 0.65
DEEP_RETRY_DELAY_SECONDS = 1.1
MIN_DEEP_COMPLETENESS = 0.68
MAX_DEEP_BULLETS = 20
JSON_MARKERS = [
    '__NEXT_DATA__',
    '__PRELOADED_STATE__',
    '__INITIAL_STATE__',
    '__WML_REDUX_INITIAL_STATE__',
    '__WML_INITIAL_STATE__',
    '__APOLLO_STATE__',
]
GENERIC_BAD_VALUE_PATTERNS = [
    r'^see more$',
    r'^view more$',
    r'^learn more$',
    r'^show more$',
    r'^read more$',
    r'^details$',
    r'^description$',
    r'^product details$',
    r'^select options?$',
    r'^choose options?$',
    r'^add to cart$',
    r'^buy now$',
    r'^more details$',
    r'^shop all$',
]

STOP_LINE_PATTERNS = [
    r"^view all item details$",
    r"^specs?$",
    r"^specifications?$",
    r"^how do you want your item\??$",
    r"^current price",
    r"^price when purchased online$",
    r"^sold by$",
    r"^fulfilled by walmart$",
    r"^free 90-day returns$",
    r"^details$",
    r"^more seller options",
    r"^about this item$",
    r"^info:$",
    r"^more details$",
    r"^customer ratings",
    r"^rating and reviews",
    r"^ratings and reviews",
    r"^product details$",
    r"^customers also considered$",
    r"^you may also like$",
    r"^similar items",
    r"^frequently bought together$",
    r"^recommended for you$",
]

CAPTCHA_PATTERNS = [
    r"verify your identity",
    r"robot or human",
    r"press and hold",
    r"access denied",
    r"captcha",
    r"are you a real person",
]

FIELD_GROUPS = [
    (
        "Core content",
        [
            ("title", "Title", True),
            ("description", "Description", True),
            ("bullet_1", "Bullet 1", True),
            ("bullet_2", "Bullet 2", True),
            ("bullet_3", "Bullet 3", True),
            ("bullet_4", "Bullet 4", True),
            ("bullet_5", "Bullet 5", True),
        ],
    ),
    (
        "Product attributes",
        [
            ("gender", "Gender", False),
            ("count_per_pack", "Count Per Pack", False),
            ("multipack", "Multipack", False),
            ("total_count", "Total Count", False),
            ("size", "Size", False),
            ("serving_size", "Serving Size", False),
            ("color", "Color", False),
            ("flavour", "Flavour", False),
            ("product_form", "Product Form", False),
        ],
    ),
    (
        "Usage and safety",
        [
            ("ingredient_statement", "Ingredient Statement", False),
            ("dosage", "Dosage", False),
            ("directions", "Directions", False),
            ("instructions", "Instructions", False),
            ("stop_use_indications", "Stop Use Indications", False),
            ("health_concern", "Health Concern", False),
        ],
    ),
    (
        "Images",
        [
            ("main_image_links", "Main Image Links", False),
            ("additional_image_1", "Additional Image 1", False),
            ("additional_image_2", "Additional Image 2", False),
            ("additional_image_3", "Additional Image 3", False),
            ("additional_image_4", "Additional Image 4", False),
            ("additional_image_5", "Additional Image 5", False),
        ],
    ),
]

FIELD_LABELS = {key: label for _, items in FIELD_GROUPS for key, label, _ in items}
FIELD_DEFAULTS = {key: default for _, items in FIELD_GROUPS for key, _, default in items}
FIELD_ORDER = [key for _, items in FIELD_GROUPS for key, _, _ in items]
LONG_TEXT_FIELDS = {
    "description",
    "ingredient_statement",
    "dosage",
    "directions",
    "instructions",
    "stop_use_indications",
}

FIELD_ALIASES = {
    "gender": ["gender"],
    "count_per_pack": ["count per pack", "countperpack", "pieces per pack", "units per pack", "item package quantity"],
    "multipack": ["multipack", "multi pack", "pack size", "number of packs", "pack quantity"],
    "total_count": ["total count", "totalcount", "count", "quantity", "item count", "unit count", "number of pieces", "number of units"],
    "size": ["size", "item size", "size unit"],
    "serving_size": ["serving size", "servingsize", "serving amount"],
    "color": ["color", "colour", "color family"],
    "flavour": ["flavor", "flavour", "taste", "scent"],
    "product_form": ["product form", "form", "item form"],
    "ingredient_statement": ["ingredient statement", "ingredients", "ingredient", "ingredient list"],
    "dosage": ["dosage", "dose"],
    "directions": ["directions", "direction", "how to use"],
    "instructions": ["instructions", "instruction", "usage instructions"],
    "stop_use_indications": ["stop use indications", "stop use indication", "warning", "warnings", "caution"],
    "health_concern": ["health concern", "health concerns", "health focus"],
}

FIELD_REGEX_PATTERNS = {
    "gender": [r'"gender"\s*:\s*"(.*?)"'],
    "count_per_pack": [r'"countPerPack"\s*:\s*"(.*?)"', r'"count_per_pack"\s*:\s*"(.*?)"'],
    "multipack": [r'"multiPack"\s*:\s*"(.*?)"', r'"multipack"\s*:\s*"(.*?)"'],
    "total_count": [r'"totalCount"\s*:\s*"(.*?)"', r'"total_count"\s*:\s*"(.*?)"'],
    "size": [r'"size"\s*:\s*"(.*?)"'],
    "serving_size": [r'"servingSize"\s*:\s*"(.*?)"', r'"serving_size"\s*:\s*"(.*?)"'],
    "color": [r'"color"\s*:\s*"(.*?)"', r'"colour"\s*:\s*"(.*?)"'],
    "flavour": [r'"flavor"\s*:\s*"(.*?)"', r'"flavour"\s*:\s*"(.*?)"', r'"taste"\s*:\s*"(.*?)"'],
    "product_form": [r'"productForm"\s*:\s*"(.*?)"', r'"form"\s*:\s*"(.*?)"'],
    "ingredient_statement": [r'"ingredientStatement"\s*:\s*"(.*?)"', r'"ingredients"\s*:\s*"(.*?)"'],
    "dosage": [r'"dosage"\s*:\s*"(.*?)"'],
    "directions": [r'"directions"\s*:\s*"(.*?)"'],
    "instructions": [r'"instructions"\s*:\s*"(.*?)"'],
    "stop_use_indications": [r'"stopUseIndications"\s*:\s*"(.*?)"', r'"warnings"\s*:\s*"(.*?)"'],
    "health_concern": [r'"healthConcern"\s*:\s*"(.*?)"', r'"healthConcern[s]?"\s*:\s*"(.*?)"'],
}

SECTION_PATTERNS = {
    "ingredient_statement": [r"^ingredient statement$", r"^ingredients?$"],
    "dosage": [r"^dosage$"],
    "directions": [r"^directions?$"],
    "instructions": [r"^instructions?$"],
    "stop_use_indications": [r"^stop use indications?$", r"^warnings?$", r"^warning$"],
}

IMAGE_PATH_ALLOW_HINTS = {
    "image", "images", "image info", "all images", "main image", "main image links",
    "mainimage", "mainimageurl", "imageurl", "image url", "imageurls", "image urls",
    "thumbnail", "thumbnails", "gallery", "media", "zoom", "zoomimage", "largeimage",
    "secondaryimage", "alternateimage", "alt image", "hero image", "heroimage", "swatchimage",
}

IMAGE_PATH_BLOCK_HINTS = {
    "logo", "icon", "sprite", "placeholder", "spinner", "badge", "rating", "review",
    "recommend", "recommended", "similar", "sponsored", "ad", "ads", "carousel", "module",
    "related", "upsell", "cross sell", "crosssell", "brand", "footer", "header", "nav",
    "swatch", "flag", "sticker", "promo", "marketing", "thumbnail icon",
}

DASHBOARD_PANEL_MIN_HEIGHT = 640
ATTRIBUTES_COLUMNS_PER_GROUP = 3

st.set_page_config(
    page_title=APP_TITLE,
    layout="wide",
    initial_sidebar_state="collapsed",
)


@dataclass
class ExtractionBundle:
    title: str = ""
    description: str = ""
    bullets: List[str] = field(default_factory=list)
    images: List[str] = field(default_factory=list)
    field_values: Dict[str, str] = field(default_factory=dict)
    error: str = ""


@st.cache_resource(show_spinner=False)
def get_requests_session() -> requests.Session:
    session = requests.Session()
    retry = Retry(
        total=2,
        connect=2,
        read=2,
        status=2,
        backoff_factor=0.8,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=frozenset({"GET", "HEAD", "OPTIONS"}),
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry, pool_connections=20, pool_maxsize=20)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    session.headers.update(REQUEST_HEADERS)
    return session


def inject_css() -> None:
    st.markdown(
        """
        <style>
            .stApp {
                background: linear-gradient(180deg, #f5f8ff 0%, #fbfcff 48%, #ffffff 100%);
                color: #0f172a;
            }
            .block-container {
                max-width: none;
                padding-top: 1rem;
                padding-bottom: 1.2rem;
            }
            .hero-card {
                background: linear-gradient(135deg, #0053e2 0%, #1d8bff 100%);
                color: white;
                border-radius: 22px;
                padding: 28px 30px;
                box-shadow: 0 22px 42px rgba(0, 83, 226, 0.18);
                margin-bottom: 1rem;
            }
            .hero-title {
                font-size: 2rem;
                font-weight: 800;
                margin: 0;
                letter-spacing: 0.2px;
            }
            .hero-subtitle {
                font-size: 1rem;
                margin-top: 0.5rem;
                opacity: 0.96;
                line-height: 1.55;
            }
            .soft-note {
                color: #5b6472;
                font-size: 0.92rem;
                line-height: 1.55;
            }
            .mini-kpi {
                background: rgba(255, 255, 255, 0.92);
                border: 1px solid rgba(0, 0, 0, 0.05);
                border-radius: 16px;
                padding: 12px 16px;
                box-shadow: 0 8px 24px rgba(15, 23, 42, 0.04);
                min-height: 84px;
            }
            .mini-kpi .label {
                font-size: 0.83rem;
                color: #5b6472;
                margin-bottom: 0.35rem;
            }
            .mini-kpi .value {
                font-size: 1.55rem;
                font-weight: 800;
                color: #0f172a;
            }
            .mini-kpi .sub {
                font-size: 0.82rem;
                color: #64748b;
                margin-top: 0.15rem;
            }
            .table-note {
                font-size: 0.88rem;
                color: #576172;
                margin-top: 0.35rem;
            }
            div[data-testid="stDownloadButton"] button,
            div[data-testid="stButton"] button {
                width: 100%;
                border-radius: 12px;
                font-weight: 700;
                padding-top: 0.65rem;
                padding-bottom: 0.65rem;
            }
            div[data-testid="stDataEditor"] {
                border-radius: 18px;
                overflow: hidden;
            }
            .status-box {
                background: #0f172a;
                color: white;
                border-radius: 16px;
                padding: 14px 16px;
            }
            .status-line {
                font-size: 0.94rem;
                line-height: 1.55;
            }
            .profile-card {
                background: rgba(255, 255, 255, 0.92);
                border: 1px solid rgba(0, 0, 0, 0.06);
                border-radius: 18px;
                padding: 16px 16px 12px 16px;
                box-shadow: 0 8px 24px rgba(15, 23, 42, 0.05);
                margin-bottom: 0.75rem;
            }
            .profile-label {
                color: #64748b;
                font-size: 0.78rem;
                margin-bottom: 0.3rem;
            }
            .profile-name {
                color: #0f172a;
                font-weight: 800;
                font-size: 1.1rem;
                margin-bottom: 0.25rem;
            }
            .profile-email {
                color: #5b6472;
                font-size: 0.82rem;
                word-break: break-word;
            }
            .login-title {
                font-size: 2rem;
                font-weight: 800;
                color: #0f172a;
                margin-bottom: 0.35rem;
            }
            .login-subtitle {
                font-size: 0.98rem;
                color: #5b6472;
                line-height: 1.5;
                margin-bottom: 0.75rem;
            }
            .login-note {
                font-size: 0.84rem;
                color: #6b7280;
                line-height: 1.55;
                margin-top: 0.75rem;
                text-align: center;
            }
            .login-footer {
                text-align: center;
                font-size: 0.8rem;
                color: #7a8597;
            }
            div[data-testid="stVerticalBlockBorderWrapper"]:has(.input-panel-anchor),
            div[data-testid="stVerticalBlockBorderWrapper"]:has(.attributes-panel-anchor) {
                height: var(--dashboard-panel-height, 640px);
            }
            div[data-testid="stVerticalBlockBorderWrapper"]:has(.input-panel-anchor) > div,
            div[data-testid="stVerticalBlockBorderWrapper"]:has(.attributes-panel-anchor) > div {
                height: 100%;
            }
            div[data-testid="stVerticalBlockBorderWrapper"]:has(.attributes-panel-anchor) {
                overflow-y: auto;
            }
        </style>
        """,
        unsafe_allow_html=True,
    )


def inject_login_css() -> None:
    st.markdown(
        """
        <style>
            :root {
                color-scheme: light;
            }
            .stApp,
            [data-testid="stAppViewContainer"],
            body {
                background: linear-gradient(180deg, #f5f8ff 0%, #fbfcff 48%, #ffffff 100%);
                color: #0f172a;
            }
            [data-testid="stHeader"] {
                background: transparent;
            }
            section.main > div.block-container {
                max-width: 1100px;
                min-height: calc(100vh - 3rem);
                display: flex;
                flex-direction: column;
                justify-content: center;
                padding-top: 0.75rem !important;
                padding-bottom: 0.9rem !important;
            }
            .login-title {
                font-size: 2rem;
                font-weight: 800;
                color: #0f172a;
                margin-bottom: 0.35rem;
            }
            .login-subtitle {
                font-size: 0.98rem;
                color: #5b6472;
                line-height: 1.5;
                margin-bottom: 0.95rem;
            }
            .login-note {
                font-size: 0.84rem;
                color: #6b7280;
                line-height: 1.55;
                margin-top: 0.9rem;
                text-align: center;
            }
            .login-footer {
                margin-top: 1rem;
                text-align: center;
                font-size: 0.8rem;
                color: #7a8597;
            }
            .login-card-frame {
                border-radius: 22px;
                border: 1px solid rgba(0, 0, 0, 0.08);
                background: rgba(255, 255, 255, 0.98);
                box-shadow: 0 18px 40px rgba(15, 23, 42, 0.08);
                padding: 0.2rem;
            }
            div[data-testid="stTextInput"] label p,
            div[data-testid="stMarkdownContainer"] p,
            div[data-testid="stMarkdownContainer"] h1,
            div[data-testid="stMarkdownContainer"] h2,
            div[data-testid="stMarkdownContainer"] h3 {
                color: #0f172a;
            }
            div[data-testid="stTextInput"] input {
                color: #0f172a !important;
                background: #f8fafc;
            }
            div[data-testid="stFormSubmitButton"] > button,
            div[data-testid="stButton"] > button {
                width: 100%;
                border-radius: 12px;
                font-weight: 700;
                padding-top: 0.72rem;
                padding-bottom: 0.72rem;
            }
        </style>
        """,
        unsafe_allow_html=True,
    )

def inject_panel_height_css(panel_height: int) -> None:
    safe_height = max(panel_height, DASHBOARD_PANEL_MIN_HEIGHT)
    st.markdown(
        f"""
        <style>
            :root {{
                --dashboard-panel-height: {safe_height}px;
            }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def init_state() -> None:
    if "input_df" not in st.session_state:
        st.session_state.input_df = build_blank_input_df(DEFAULT_ROWS)
    if "results_df" not in st.session_state:
        st.session_state.results_df = pd.DataFrame(columns=default_output_columns())
    if "last_failures" not in st.session_state:
        st.session_state.last_failures = []
    if "uploaded_signature" not in st.session_state:
        st.session_state.uploaded_signature = None
    if "file_name" not in st.session_state:
        st.session_state.file_name = DEFAULT_OUTPUT_FILENAME
    if "row_count" not in st.session_state:
        st.session_state.row_count = DEFAULT_ROWS
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    if "user_email" not in st.session_state:
        st.session_state.user_email = ""
    if "user_name" not in st.session_state:
        st.session_state.user_name = ""
    for field_key in FIELD_ORDER:
        state_key = f"sel_{field_key}"
        if state_key not in st.session_state:
            st.session_state[state_key] = FIELD_DEFAULTS[field_key]
    if "expand_extra_bullets" not in st.session_state:
        st.session_state.expand_extra_bullets = False
    if "expand_extra_images" not in st.session_state:
        st.session_state.expand_extra_images = False


def build_blank_input_df(rows: int) -> pd.DataFrame:
    rows = max(1, min(int(rows), MAX_ROWS))
    return pd.DataFrame({col: [""] * rows for col in INPUT_COLUMNS})


def ensure_row_count(df: pd.DataFrame, rows: int) -> pd.DataFrame:
    rows = max(1, min(int(rows), MAX_ROWS))
    base = df.copy()
    for col in INPUT_COLUMNS:
        if col not in base.columns:
            base[col] = ""
    base = base[INPUT_COLUMNS]
    if len(base) > rows:
        base = base.iloc[:rows].copy()
    elif len(base) < rows:
        extra = pd.DataFrame({col: [""] * (rows - len(base)) for col in INPUT_COLUMNS})
        base = pd.concat([base, extra], ignore_index=True)
    return base.fillna("")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\u00a0", " ")
    text = text.replace("\u200b", "")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def clean_multiline_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ").replace("\u200b", "")
    text = re.sub(r"\r\n?|\n", "\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def coerce_input_df(df: pd.DataFrame) -> pd.DataFrame:
    base = df.copy()
    for col in INPUT_COLUMNS:
        if col not in base.columns:
            base[col] = ""
        base[col] = base[col].map(clean_text)
    return base[INPUT_COLUMNS]


def normalize_url(url: str) -> str:
    value = clean_text(url)
    if not value:
        return ""
    if value.startswith("www."):
        value = f"https://{value}"
    elif not re.match(r"^https?://", value, flags=re.I):
        value = f"https://{value}"
    return value


def looks_like_walmart_url(url: str) -> bool:
    return bool(re.search(r"https?://([a-z0-9-]+\.)?walmart\.com/", url, flags=re.I))


def slugify_filename(name: str) -> str:
    cleaned = clean_text(name) or DEFAULT_OUTPUT_FILENAME
    cleaned = re.sub(r"[\\/:*?\"<>|]+", "_", cleaned)
    cleaned = re.sub(r"\s+", "_", cleaned)
    cleaned = cleaned.strip("._")
    return cleaned[:120] or DEFAULT_OUTPUT_FILENAME


def first_non_empty(values: Iterable[Optional[str]]) -> str:
    for value in values:
        text = clean_text(value)
        if text:
            return text
    return ""


def dedupe_keep_order(items: Iterable[str]) -> List[str]:
    seen: set[str] = set()
    output: List[str] = []
    for item in items:
        text = clean_text(item)
        key = text.lower()
        if text and key not in seen:
            seen.add(key)
            output.append(text)
    return output


def parse_uploaded_dataframe(uploaded_file) -> Tuple[pd.DataFrame, str]:
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        df = pd.read_csv(uploaded_file, dtype=str, keep_default_na=False)
    else:
        df = pd.read_excel(uploaded_file, dtype=str)

    if df.empty:
        return build_blank_input_df(DEFAULT_ROWS), "Uploaded file is empty."

    normalized = {str(col).strip().lower(): col for col in df.columns}
    sku_col = None
    url_col = None
    for key, original in normalized.items():
        if sku_col is None and "sku" in key:
            sku_col = original
        if url_col is None and ("url" in key or "walmart" in key or "listing" in key):
            url_col = original

    if sku_col is None and len(df.columns) >= 1:
        sku_col = df.columns[0]
    if url_col is None and len(df.columns) >= 2:
        url_col = df.columns[1]

    if sku_col is None or url_col is None:
        raise ValueError("Could not detect SKU and URL columns in the uploaded file.")

    parsed = pd.DataFrame(
        {
            "SKU": df[sku_col].map(clean_text),
            "Walmart URL": df[url_col].map(clean_text),
        }
    )
    parsed = parsed.iloc[:MAX_ROWS].copy().fillna("")
    message = f"Loaded {len(parsed)} rows from the uploaded file."
    if len(df) > MAX_ROWS:
        message += f" Only the first {MAX_ROWS} rows were kept."
    return parsed, message


def default_output_columns() -> List[str]:
    columns = ["SKU", "URL"]
    for key in FIELD_ORDER:
        if FIELD_DEFAULTS[key]:
            columns.append(FIELD_LABELS[key])
    return columns


def normalize_label(text: str) -> str:
    value = clean_text(text).lower()
    value = value.replace("&", " and ")
    value = re.sub(r"[_/]+", " ", value)
    value = re.sub(r"[^a-z0-9]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def line_matches_any(line: str, patterns: Iterable[str]) -> bool:
    text = clean_text(line)
    return any(re.search(pattern, text, flags=re.I) for pattern in patterns)


def detect_captcha_or_block(lines: List[str], html: str = "") -> bool:
    joined = " ".join(lines[:120]).lower()
    combined = f"{joined} {html[:5000].lower()}"
    return any(re.search(pattern, combined, flags=re.I) for pattern in CAPTCHA_PATTERNS)


def strip_bullet_prefix(line: str) -> str:
    text = clean_text(line)
    text = re.sub(r"^[\u2022\-*\s]+", "", text)
    return text


def is_generic_section_heading(line: str) -> bool:
    return line_matches_any(
        line,
        [
            r"^key item features$",
            r"^highlights$",
            r"^about this item$",
            r"^product details$",
            r"^description$",
            r"^details$",
            r"^specs?$",
            r"^specifications?$",
            r"^more details$",
        ],
    )


def is_short_label_fragment(line: str) -> bool:
    text = strip_bullet_prefix(line)
    if not text or len(text) > 45:
        return False
    if line_matches_any(text, STOP_LINE_PATTERNS) or is_generic_section_heading(text):
        return False
    if re.search(r"[.!?]$", text):
        return False
    words = re.findall(r"[A-Za-z0-9&'+/\-]+", text)
    return 1 <= len(words) <= 6


def merge_fragmented_lines(lines: List[str]) -> List[str]:
    merged: List[str] = []
    idx = 0
    while idx < len(lines):
        line = clean_text(lines[idx])
        if not line:
            idx += 1
            continue
        if line.startswith(":") and merged and is_short_label_fragment(merged[-1]):
            merged[-1] = clean_text(f"{merged[-1]}{line}")
            idx += 1
            continue
        next_line = clean_text(lines[idx + 1]) if idx + 1 < len(lines) else ""
        if next_line:
            if is_short_label_fragment(line) and next_line.startswith(":"):
                merged.append(clean_text(f"{line}{next_line}"))
                idx += 2
                continue
            if line.endswith(":") and not line_matches_any(next_line, STOP_LINE_PATTERNS):
                merged.append(clean_text(f"{line} {strip_bullet_prefix(next_line)}"))
                idx += 2
                continue
        merged.append(line)
        idx += 1
    return merged


def parse_visible_lines(body_text: str) -> List[str]:
    output: List[str] = []
    for raw in body_text.splitlines():
        line = clean_text(raw)
        if not line:
            continue
        output.append(line)
    return merge_fragmented_lines(output)


def looks_like_bullet_line(line: str) -> bool:
    text = strip_bullet_prefix(line)
    if len(text) < 20:
        return False
    if text != clean_text(line):
        return True
    if ":" in text[:55]:
        head = text.split(":", 1)[0].strip()
        letters = re.sub(r"[^A-Za-z]", "", head)
        if letters and (
            head.isupper() or sum(ch.isupper() for ch in head) >= max(3, int(len(letters) * 0.45))
        ):
            return True
    return False


def find_heading_index(lines: List[str], patterns: Iterable[str]) -> Optional[int]:
    for idx, line in enumerate(lines):
        if line_matches_any(line, patterns):
            return idx
    return None


def collect_section_lines(lines: List[str], start_idx: int, max_scan_lines: int = 45) -> List[str]:
    section: List[str] = []
    for line in lines[start_idx + 1 : start_idx + 1 + max_scan_lines]:
        if line_matches_any(line, STOP_LINE_PATTERNS):
            break
        if re.search(r"accurate product information|see our disclaimer", line, flags=re.I):
            break
        if is_generic_section_heading(line) and section:
            break
        section.append(line)
    return merge_fragmented_lines(section)


def extract_section_text(lines: List[str], heading_patterns: Iterable[str], max_scan_lines: int = 24) -> str:
    start_idx = find_heading_index(lines, heading_patterns)
    if start_idx is None:
        return ""
    pieces: List[str] = []
    for line in collect_section_lines(lines, start_idx, max_scan_lines=max_scan_lines):
        text = strip_bullet_prefix(line)
        if not text:
            continue
        if pieces and is_short_label_fragment(text) and not text.endswith(":"):
            break
        pieces.append(text)
        if len(" ".join(pieces)) >= 1200:
            break
    return clean_multiline_text("\n".join(dedupe_keep_order(pieces)))


def extract_description_from_lines(lines: List[str]) -> str:
    start_idx = find_heading_index(lines, [r"^product details$", r"^description$"])
    if start_idx is None:
        return ""
    pieces: List[str] = []
    for line in collect_section_lines(lines, start_idx, max_scan_lines=40):
        if looks_like_bullet_line(line) or line.startswith(":"):
            break
        if len(line) < 25:
            if pieces:
                break
            continue
        pieces.append(strip_bullet_prefix(line))
        if len(" ".join(pieces)) >= 1200 or len(pieces) >= 5:
            break
    return clean_text(" ".join(dedupe_keep_order(pieces)))


def extract_bullets_from_lines(lines: List[str]) -> List[str]:
    bullets: List[str] = []
    key_idx = find_heading_index(lines, [r"^key item features$", r"^highlights$", r"^about this item$"])
    if key_idx is not None:
        for line in collect_section_lines(lines, key_idx, max_scan_lines=50):
            candidate = strip_bullet_prefix(line)
            if len(candidate) < 12:
                continue
            bullets.append(candidate)
            if len(bullets) >= 20:
                break
    if bullets:
        return dedupe_keep_order(bullets)[:20]

    detail_idx = find_heading_index(lines, [r"^product details$", r"^description$"])
    if detail_idx is not None:
        description_started = False
        for line in collect_section_lines(lines, detail_idx, max_scan_lines=50):
            candidate = strip_bullet_prefix(line)
            if looks_like_bullet_line(candidate) or (":" in candidate[:80] and len(candidate) >= 20):
                description_started = True
                bullets.append(candidate)
            elif description_started:
                break
            if len(bullets) >= 20:
                break
    return dedupe_keep_order(bullets)[:20]


def normalize_json_text(value: str) -> str:
    text = clean_text(value)
    if not text:
        return ""
    text = unescape(text)
    text = text.replace('\u003c', '<').replace('\u003e', '>').replace('\u0026', '&')
    try:
        text = bytes(text, 'utf-8').decode('unicode_escape')
    except Exception:
        pass
    return clean_multiline_text(text)

def canonicalize_url(url: str) -> str:
    text = clean_text(url)
    text = text.replace('\\/', '/')
    text = text.replace('\\u002F', '/')
    text = text.replace('&amp;', '&')
    if text.startswith('//'):
        text = f'https:{text}'
    if text.startswith('http://'):
        text = 'https://' + text[len('http://'):]
    return text


def is_probable_image_url(value: str, allow_png: bool = False) -> bool:
    text = clean_text(value)
    if not text:
        return False
    text = canonicalize_url(text)
    text_lower = text.lower()
    if text_lower.startswith('data:') or text_lower.endswith('.svg'):
        return False
    if any(token in text_lower for token in ['logo', 'icon', 'sprite', 'placeholder', 'spinner', 'badge', 'rating']):
        return False
    extension_match = re.search(r'\.([a-z0-9]{3,4})(?:\?|$)', text_lower)
    extension = extension_match.group(1) if extension_match else ''
    if extension == 'png' and not allow_png:
        return False
    allowed_extensions = {'jpg', 'jpeg', 'webp'} | ({'png'} if allow_png else set())
    if extension and extension not in allowed_extensions:
        return False
    if 'walmartimages' in text_lower:
        return True
    return bool(re.search(r'\.(jpg|jpeg|webp)(\?|$)', text_lower)) or (allow_png and bool(re.search(r'\.png(\?|$)', text_lower)))


def path_looks_like_product_image(path_parts: Iterable[str]) -> bool:
    normalized_parts = [normalize_label(part) for part in path_parts if normalize_label(part)]
    if not normalized_parts:
        return False
    path_text = ' '.join(normalized_parts)
    if any(block in path_text for block in IMAGE_PATH_BLOCK_HINTS):
        return False
    return any(allow in path_text for allow in IMAGE_PATH_ALLOW_HINTS)


def add_candidate(candidate_map: Dict[str, List[str]], label: str, value: str) -> None:
    normalized_label = normalize_label(label)
    cleaned_value = clean_multiline_text(value)
    if not normalized_label or not cleaned_value:
        return
    existing = candidate_map.setdefault(normalized_label, [])
    if cleaned_value not in existing:
        existing.append(cleaned_value)


def walk_json_object(
    obj: Any,
    candidate_map: Dict[str, List[str]],
    image_urls: List[str],
    current_key: Optional[str] = None,
    path_parts: Optional[List[str]] = None,
) -> None:
    path_parts = path_parts or []
    if isinstance(obj, dict):
        for key, value in obj.items():
            normalized_key = normalize_label(str(key))
            walk_json_object(
                value,
                candidate_map,
                image_urls,
                current_key=normalized_key,
                path_parts=path_parts + [normalized_key],
            )
        return
    if isinstance(obj, list):
        for item in obj:
            walk_json_object(item, candidate_map, image_urls, current_key=current_key, path_parts=path_parts)
        return
    if obj is None:
        return
    text = normalize_json_text(str(obj))
    if not text:
        return
    if current_key:
        add_candidate(candidate_map, current_key, text)
    if is_probable_image_url(text) and path_looks_like_product_image(path_parts):
        image_urls.append(canonicalize_url(text))



def looks_like_bad_value(value: str) -> bool:
    text = clean_text(value)
    if not text:
        return True
    lowered = text.lower()
    if any(re.search(pattern, lowered, flags=re.I) for pattern in GENERIC_BAD_VALUE_PATTERNS):
        return True
    if line_matches_any(text, STOP_LINE_PATTERNS):
        return True
    return False


def extract_balanced_json_fragment(raw: str, marker: str) -> List[str]:
    fragments: List[str] = []
    search_from = 0
    while True:
        marker_idx = raw.find(marker, search_from)
        if marker_idx == -1:
            break
        eq_idx = raw.find('=', marker_idx)
        if eq_idx == -1:
            search_from = marker_idx + len(marker)
            continue
        start = -1
        for pos in range(eq_idx + 1, min(len(raw), eq_idx + 400)):
            if raw[pos] in '{[':
                start = pos
                break
        if start == -1:
            search_from = eq_idx + 1
            continue
        opening = raw[start]
        closing = '}' if opening == '{' else ']'
        depth = 0
        in_string = False
        escape = False
        for pos in range(start, len(raw)):
            ch = raw[pos]
            if in_string:
                if escape:
                    escape = False
                elif ch == '\\':
                    escape = True
                elif ch == '"':
                    in_string = False
                continue
            if ch == '"':
                in_string = True
                continue
            if ch == opening:
                depth += 1
            elif ch == closing:
                depth -= 1
                if depth == 0:
                    fragments.append(raw[start:pos + 1])
                    search_from = pos + 1
                    break
        else:
            search_from = start + 1
    return fragments


def iter_json_payloads_from_script(script: Any) -> Iterable[Any]:
    raw = (script.string or script.get_text() or '').strip()
    if not raw:
        return []
    payloads: List[Any] = []
    if raw.startswith('{') or raw.startswith('['):
        try:
            payloads.append(json.loads(raw))
        except Exception:
            pass
    else:
        for marker in JSON_MARKERS:
            if marker not in raw:
                continue
            for fragment in extract_balanced_json_fragment(raw, marker):
                try:
                    payloads.append(json.loads(fragment))
                except Exception:
                    continue
    return payloads


def candidate_key_match_score(key: str, alias: str) -> int:
    if not key or not alias:
        return 0
    if key == alias:
        return 100
    score = 0
    if alias in key or key in alias:
        score = max(score, 84)
    key_tokens = set(key.split())
    alias_tokens = set(alias.split())
    overlap = key_tokens & alias_tokens
    if overlap:
        score = max(score, 60 + len(overlap) * 10 - abs(len(key_tokens) - len(alias_tokens)) * 2)
    if key.startswith(alias + ' ') or alias.startswith(key + ' '):
        score = max(score, 88)
    return score


def extract_description_from_soup(soup: BeautifulSoup) -> str:
    selectors = [
        '[data-testid*="description"]',
        '[itemprop="description"]',
        '#product-description',
        '[aria-label*="description" i]',
        '[data-automation-id*="description"]',
    ]
    candidates: List[str] = []
    for selector in selectors:
        for node in soup.select(selector):
            text = clean_multiline_text(node.get_text(' ', strip=True))
            if len(text) >= 70 and not looks_like_bad_value(text):
                candidates.append(text)
    return max(candidates, key=len) if candidates else ''


def split_bullet_candidate_text(text: str) -> List[str]:
    candidate = clean_multiline_text(text)
    if not candidate or looks_like_bad_value(candidate):
        return []
    pieces = [candidate]
    if '\n' in candidate:
        pieces = [part.strip() for part in re.split(r'\n+', candidate)]
    elif any(token in candidate for token in ['•', '●', '|']):
        pieces = [part.strip() for part in re.split(r'\s*[•●|]\s*', candidate)]
    elif candidate.count(';') >= 2 and len(candidate) >= 80:
        pieces = [part.strip() for part in re.split(r'\s*;\s*', candidate)]
    elif re.search(r'\b1\.\s+\S+', candidate):
        pieces = [part.strip() for part in re.split(r'(?=\b\d+\.\s+)', candidate) if part.strip()]
    cleaned: List[str] = []
    for piece in pieces:
        item = strip_bullet_prefix(piece)
        item = re.sub(r'^\d+\.\s*', '', item)
        item = clean_text(item)
        if len(item) < 12 or looks_like_bad_value(item):
            continue
        cleaned.append(item)
    return dedupe_keep_order(cleaned)


def extract_bullets_from_candidate_maps(candidate_maps: List[Dict[str, List[str]]]) -> List[str]:
    bullets: List[str] = []
    bullet_aliases = [
        'key item features', 'highlights', 'features', 'feature', 'bullets', 'bullet',
        'about this item', 'product highlights', 'key features', 'important information',
    ]
    for candidate_map in candidate_maps:
        for key, values in candidate_map.items():
            best_score = max((candidate_key_match_score(key, alias) for alias in bullet_aliases), default=0)
            if best_score < 72:
                continue
            for value in values:
                for item in split_bullet_candidate_text(value):
                    bullets.append(item)
                    if len(bullets) >= MAX_DEEP_BULLETS:
                        return dedupe_keep_order(bullets)[:MAX_DEEP_BULLETS]
    return dedupe_keep_order(bullets)[:MAX_DEEP_BULLETS]


def is_good_field_value(field_key: str, value: str) -> bool:
    text = clean_multiline_text(value)
    if not text or looks_like_bad_value(text):
        return False
    if field_key not in LONG_TEXT_FIELDS and len(text) > 500:
        return False
    if field_key in {'gender', 'color', 'flavour', 'product_form'} and len(text.split()) > 12:
        return False
    return True


def compute_bundle_completeness(bundle: ExtractionBundle, selected_field_keys: List[str]) -> float:
    if not selected_field_keys:
        return 1.0
    found = 0
    for key in selected_field_keys:
        if key == 'title' and bundle.title:
            found += 1
        elif key == 'description' and bundle.description:
            found += 1
        elif key.startswith('bullet_'):
            try:
                number = int(key.split('_')[1])
            except Exception:
                number = 0
            if number and len(bundle.bullets or []) >= number:
                found += 1
        elif key == 'main_image_links' and bundle.images:
            found += 1
        elif key.startswith('additional_image_'):
            try:
                number = int(key.rsplit('_', 1)[1])
            except Exception:
                number = 0
            if number and len(bundle.images or []) >= number + 1:
                found += 1
        elif bundle.field_values.get(key):
            found += 1
    return found / max(len(selected_field_keys), 1)


def merge_bundles(primary: ExtractionBundle, secondary: ExtractionBundle) -> ExtractionBundle:
    merged_fields = dict(primary.field_values)
    for key, value in secondary.field_values.items():
        if not merged_fields.get(key):
            merged_fields[key] = value
        elif len(clean_multiline_text(value)) > len(clean_multiline_text(merged_fields[key])):
            merged_fields[key] = value
    return ExtractionBundle(
        title=primary.title if len(primary.title) >= len(secondary.title) else secondary.title,
        description=primary.description if len(primary.description) >= len(secondary.description) else secondary.description,
        bullets=dedupe_keep_order((primary.bullets or []) + (secondary.bullets or []))[:MAX_DEEP_BULLETS],
        images=dedupe_keep_order((primary.images or []) + (secondary.images or [])),
        field_values=merged_fields,
        error=primary.error or secondary.error,
    )


def extract_json_candidates_from_soup(soup: BeautifulSoup) -> Tuple[Dict[str, List[str]], List[str]]:
    candidate_map: Dict[str, List[str]] = {}
    image_urls: List[str] = []
    for script in soup.find_all('script'):
        for parsed in iter_json_payloads_from_script(script):
            try:
                walk_json_object(parsed, candidate_map, image_urls)
            except Exception:
                continue
    return candidate_map, dedupe_keep_order(image_urls)

def extract_embedded_text_candidates(html: str) -> Dict[str, str]:
    candidates: Dict[str, str] = {}
    patterns = {
        'description': [
            r'"shortDescription"\s*:\s*"(.*?)"',
            r'"description"\s*:\s*"(.*?)"',
            r'"productDescription"\s*:\s*"(.*?)"',
        ],
        'title': [
            r'"productName"\s*:\s*"(.*?)"',
            r'"name"\s*:\s*"(.*?)"',
        ],
    }
    for key, regexes in patterns.items():
        for pattern in regexes:
            match = re.search(pattern, html, flags=re.I | re.S)
            if match:
                value = normalize_json_text(match.group(1))
                if value:
                    candidates[key] = value
                    break
    return candidates


def clean_soup_for_text(soup: BeautifulSoup) -> BeautifulSoup:
    cloned = BeautifulSoup(str(soup), 'lxml')
    for tag in cloned(['script', 'style', 'noscript', 'svg', 'template']):
        tag.decompose()
    return cloned


def looks_like_label_text(text: str) -> bool:
    value = clean_text(text)
    if not value or len(value) > 45 or is_generic_section_heading(value):
        return False
    if line_matches_any(value, STOP_LINE_PATTERNS):
        return False
    if re.search(r'[.!?]$', value):
        return False
    if re.search(r'https?://', value, flags=re.I):
        return False
    words = re.findall(r"[A-Za-z0-9&'+/\-]+", value)
    if not (1 <= len(words) <= 7):
        return False
    letters = re.sub(r'[^A-Za-z]', '', value)
    if letters and value == value.lower() and len(words) > 3:
        return False
    return True


def build_line_candidate_map(lines: List[str]) -> Dict[str, List[str]]:
    candidate_map: Dict[str, List[str]] = {}
    total = len(lines)
    for idx, line in enumerate(lines):
        text = clean_text(line)
        if not text:
            continue
        if ':' in text and len(text.split(':', 1)[0]) <= 50:
            label, value = text.split(':', 1)
            if looks_like_label_text(label) and clean_text(value):
                add_candidate(candidate_map, label, value)
        if not looks_like_label_text(text):
            continue
        values: List[str] = []
        j = idx + 1
        while j < total and len(values) < 4:
            nxt = clean_text(lines[j])
            if not nxt or line_matches_any(nxt, STOP_LINE_PATTERNS):
                break
            if is_generic_section_heading(nxt) and values:
                break
            if looks_like_label_text(nxt) and values:
                break
            values.append(strip_bullet_prefix(nxt))
            if len(' '.join(values)) >= 300:
                break
            j += 1
        if values:
            add_candidate(candidate_map, text, ' '.join(values))
    return candidate_map


def build_dom_candidate_map(soup: BeautifulSoup) -> Dict[str, List[str]]:
    candidate_map: Dict[str, List[str]] = {}
    for row in soup.select('tr'):
        cells = [clean_text(cell.get_text(' ', strip=True)) for cell in row.find_all(['th', 'td'])]
        cells = [cell for cell in cells if cell]
        if len(cells) >= 2 and looks_like_label_text(cells[0]):
            add_candidate(candidate_map, cells[0], ' '.join(cells[1:]))
    for dt in soup.select('dt'):
        dd = dt.find_next_sibling('dd')
        if dd:
            label = clean_text(dt.get_text(' ', strip=True))
            value = clean_text(dd.get_text(' ', strip=True))
            if looks_like_label_text(label) and value:
                add_candidate(candidate_map, label, value)
    return candidate_map


def lookup_candidate(candidate_maps: Iterable[Dict[str, List[str]]], aliases: Iterable[str], prefer_long: bool = False) -> str:
    normalized_aliases = [normalize_label(alias) for alias in aliases if normalize_label(alias)]
    best_score = 0
    best_value = ''
    for candidate_map in candidate_maps:
        for key, values in candidate_map.items():
            for alias in normalized_aliases:
                score = candidate_key_match_score(key, alias)
                if score < 60:
                    continue
                usable_values = [value for value in values if is_good_field_value(alias, value)]
                if not usable_values:
                    continue
                value = max(usable_values, key=len) if prefer_long else usable_values[0]
                if score > best_score or (score == best_score and len(clean_multiline_text(value)) > len(clean_multiline_text(best_value))):
                    best_score = score
                    best_value = value
    return best_value

def regex_extract_first(html: str, patterns: Iterable[str]) -> str:
    for pattern in patterns:
        match = re.search(pattern, html, flags=re.I | re.S)
        if match:
            value = normalize_json_text(match.group(1))
            if value:
                return value
    return ''


def infer_multipack_from_text(text: str) -> str:
    candidate = clean_text(text)
    patterns = [
        r'\b(\d+)\s*pack\b',
        r'\bpack of\s*(\d+)\b',
        r'\((\d+)\s*pack\)',
    ]
    for pattern in patterns:
        match = re.search(pattern, candidate, flags=re.I)
        if match:
            return f"{match.group(1)} Pack"
    return ''


def infer_total_count_from_text(text: str) -> str:
    candidate = clean_text(text)
    match = re.search(
        r'\b(\d+(?:\.\d+)?)\s*(count|ct|gummies|tablets|capsules|softgels|chews|pieces)\b',
        candidate,
        flags=re.I,
    )
    if match:
        return clean_text(match.group(0))
    return ''


def extract_image_links(soup: BeautifulSoup, html: str, json_image_urls: List[str]) -> List[str]:
    preferred_urls: List[str] = []
    fallback_urls: List[str] = []

    for url in json_image_urls:
        if is_probable_image_url(url):
            preferred_urls.append(canonicalize_url(url))

    for attrs in [
        {'property': 'og:image'},
        {'name': 'og:image'},
        {'property': 'twitter:image'},
        {'name': 'twitter:image'},
        {'name': 'image'},
    ]:
        tag = soup.find('meta', attrs=attrs)
        if tag and tag.get('content'):
            url = canonicalize_url(tag.get('content', ''))
            if is_probable_image_url(url):
                fallback_urls.append(url)

    gallery_selectors = [
        '[data-testid*="media"] img',
        '[data-testid*="carousel"] img',
        '[data-testid*="hero"] img',
        '[data-testid*="image"] img',
        'button img[src*="walmartimages"]',
    ]
    for selector in gallery_selectors:
        for tag in soup.select(selector):
            alt_text = clean_text(tag.get('alt', ''))
            raw = tag.get('data-src') or tag.get('src') or tag.get('data-image-src')
            if not raw:
                continue
            url = canonicalize_url(raw)
            if not is_probable_image_url(url):
                continue
            if alt_text and any(token in alt_text.lower() for token in ['logo', 'icon', 'rating', 'sponsored']):
                continue
            fallback_urls.append(url)

    # Extremely broad URL regexes are intentionally avoided here because they pull in
    # logos, recommendation modules, and non-product assets from the PDP.

    return dedupe_keep_order(preferred_urls + fallback_urls)


def extract_title_from_sources(
    soup: BeautifulSoup,
    html_candidates: Dict[str, str],
    json_candidates: Dict[str, List[str]],
) -> str:
    h1 = soup.find('h1')
    og = soup.find('meta', attrs={'property': 'og:title'})
    meta_title = soup.find('meta', attrs={'name': 'title'})
    page_title = soup.find('title')
    title = first_non_empty(
        [
            h1.get_text(' ', strip=True) if h1 else '',
            og.get('content', '') if og else '',
            meta_title.get('content', '') if meta_title else '',
            page_title.get_text(' ', strip=True) if page_title else '',
            lookup_candidate([json_candidates], ['name', 'product name']),
            html_candidates.get('title', ''),
        ]
    )
    title = re.sub(r'\s*-\s*Walmart\.com\s*$', '', title, flags=re.I)
    return clean_text(title)


def extract_meta_description(soup: BeautifulSoup) -> str:
    tag = soup.find('meta', attrs={'name': 'description'})
    if not tag:
        tag = soup.find('meta', attrs={'property': 'og:description'})
    if not tag:
        return ''
    return clean_text(tag.get('content', ''))


def extract_field_value(
    field_key: str,
    lines: List[str],
    candidate_maps: List[Dict[str, List[str]]],
    html: str,
    title: str,
    description: str,
) -> str:
    aliases = FIELD_ALIASES.get(field_key, [FIELD_LABELS.get(field_key, field_key)])
    section_text = ''
    if field_key in SECTION_PATTERNS:
        section_text = extract_section_text(lines, SECTION_PATTERNS[field_key], max_scan_lines=22)
    regex_value = regex_extract_first(html, FIELD_REGEX_PATTERNS.get(field_key, []))
    candidate_value = lookup_candidate(candidate_maps, aliases, prefer_long=field_key in LONG_TEXT_FIELDS)
    value = first_non_empty([section_text, candidate_value, regex_value])
    if not value and field_key == 'multipack':
        value = infer_multipack_from_text(title)
    if not value and field_key == 'total_count':
        value = infer_total_count_from_text(title) or infer_total_count_from_text(description)
    cleaned = clean_multiline_text(value)
    return cleaned if is_good_field_value(field_key, cleaned) else ''

def scrape_listing_from_html(html: str, selected_field_keys: List[str], resolved_url: str = '') -> ExtractionBundle:
    if not clean_text(html):
        return ExtractionBundle(error='Empty page received.')

    soup = BeautifulSoup(html, 'lxml')
    html_candidates = extract_embedded_text_candidates(html)
    json_candidates, json_image_urls = extract_json_candidates_from_soup(soup)
    text_soup = clean_soup_for_text(soup)
    lines = parse_visible_lines(text_soup.get_text('\n'))

    if detect_captcha_or_block(lines, html):
        return ExtractionBundle(error='Walmart presented an anti-bot or verification page.')

    dom_candidates = build_dom_candidate_map(soup)
    line_candidates = build_line_candidate_map(lines)
    candidate_maps = [dom_candidates, line_candidates, json_candidates]

    title = ''
    description = ''
    bullets: List[str] = []
    images: List[str] = []
    field_values: Dict[str, str] = {}

    need_title = 'title' in selected_field_keys
    need_description = 'description' in selected_field_keys
    need_bullets = any(key.startswith('bullet_') for key in selected_field_keys)
    need_images = 'main_image_links' in selected_field_keys or any(key.startswith('additional_image_') for key in selected_field_keys)
    need_scalar_fields = [
        key for key in selected_field_keys
        if key not in {'title', 'description', 'main_image_links'} and not key.startswith('bullet_') and not key.startswith('additional_image_')
    ]

    if need_title or need_scalar_fields:
        title = extract_title_from_sources(soup, html_candidates, json_candidates)
    if need_description or need_scalar_fields:
        description = first_non_empty(
            [
                extract_description_from_lines(lines),
                extract_description_from_soup(soup),
                lookup_candidate([json_candidates], ['long description', 'product description', 'short description', 'description html', 'description'], prefer_long=True),
                html_candidates.get('description', ''),
                extract_meta_description(soup),
            ]
        )
        description = clean_multiline_text(description)
    if need_bullets:
        bullets = dedupe_keep_order(
            extract_bullets_from_candidate_maps(candidate_maps) + extract_bullets_from_lines(lines)
        )[:MAX_DEEP_BULLETS]
    if need_images:
        images = extract_image_links(soup, html, json_image_urls)
    for field_key in need_scalar_fields:
        field_values[field_key] = extract_field_value(field_key, lines, candidate_maps, html, title, description)

    data_found = []
    if need_title:
        data_found.append(bool(title))
    if need_description:
        data_found.append(bool(description))
    if need_bullets:
        data_found.append(bool(bullets))
    if need_images:
        data_found.append(bool(images))
    for field_key in need_scalar_fields:
        data_found.append(bool(field_values.get(field_key, '')))

    if selected_field_keys and not any(data_found):
        if resolved_url and not re.search(r'/ip/', resolved_url):
            return ExtractionBundle(error='The URL did not resolve to a Walmart product page.')
        return ExtractionBundle(error='Could not extract the selected attributes from the page.')

    return ExtractionBundle(
        title=title,
        description=description,
        bullets=bullets,
        images=images,
        field_values=field_values,
    )

def fetch_listing_html(url: str) -> Tuple[str, str]:
    session = get_requests_session()
    try:
        response = session.get(url, timeout=(15, 35), allow_redirects=True)
    except requests.RequestException as exc:
        raise RuntimeError(f'Network request failed: {exc}') from exc
    if response.status_code in {403, 429}:
        raise RuntimeError('Walmart blocked the request or rate-limited the session.')
    if response.status_code == 404:
        raise RuntimeError('Listing page returned 404 Not Found.')
    if response.status_code >= 400:
        raise RuntimeError(f'Listing page returned HTTP {response.status_code}.')
    content_type = response.headers.get('content-type', '')
    if 'html' not in content_type.lower():
        raise RuntimeError(f"Unexpected response type: {content_type or 'unknown'}")
    return response.text, response.url


def scrape_listing(url: str, selected_field_keys: List[str]) -> ExtractionBundle:
    try:
        html, resolved_url = fetch_listing_html(url)
    except Exception as exc:
        return ExtractionBundle(error=str(exc))

    primary = scrape_listing_from_html(html, selected_field_keys, resolved_url=resolved_url)
    if primary.error:
        return primary

    primary_score = compute_bundle_completeness(primary, selected_field_keys)
    if primary_score >= MIN_DEEP_COMPLETENESS:
        return primary

    time.sleep(DEEP_RETRY_DELAY_SECONDS)
    try:
        retry_html, retry_resolved_url = fetch_listing_html(url)
    except Exception:
        return primary
    secondary = scrape_listing_from_html(retry_html, selected_field_keys, resolved_url=retry_resolved_url)
    if secondary.error:
        return primary
    merged = merge_bundles(primary, secondary)
    return merged if compute_bundle_completeness(merged, selected_field_keys) >= primary_score else primary

def selected_field_keys_from_state() -> List[str]:
    return [key for key in FIELD_ORDER if bool(st.session_state.get(f'sel_{key}', False))]


def selected_bullet_numbers(selected_field_keys: List[str]) -> List[int]:
    numbers: List[int] = []
    for key in selected_field_keys:
        match = re.fullmatch(r'bullet_(\d+)', key)
        if match:
            numbers.append(int(match.group(1)))
    return sorted(numbers)


def selected_additional_image_numbers(selected_field_keys: List[str]) -> List[int]:
    numbers: List[int] = []
    for key in selected_field_keys:
        match = re.fullmatch(r'additional_image_(\d+)', key)
        if match:
            numbers.append(int(match.group(1)))
    return sorted(numbers)


def build_output_columns(
    selected_field_keys: List[str],
    max_bullet_count: int,
    max_additional_image_count: int,
    expand_extra_bullets: bool,
    expand_extra_images: bool,
) -> List[str]:
    columns = ['SKU', 'URL']
    if 'title' in selected_field_keys:
        columns.append('Title')
    if 'description' in selected_field_keys:
        columns.append('Description')

    bullet_numbers = selected_bullet_numbers(selected_field_keys)
    for number in bullet_numbers:
        columns.append(f'Bullet {number}')
    if expand_extra_bullets and bullet_numbers:
        start = max(max(bullet_numbers), 5) + 1
        for number in range(start, max_bullet_count + 1):
            columns.append(f'Bullet {number}')

    for key in [
        'gender',
        'count_per_pack',
        'multipack',
        'total_count',
        'size',
        'serving_size',
        'color',
        'flavour',
        'product_form',
        'ingredient_statement',
        'dosage',
        'directions',
        'instructions',
        'stop_use_indications',
        'health_concern',
    ]:
        if key in selected_field_keys:
            columns.append(FIELD_LABELS[key])

    if 'main_image_links' in selected_field_keys:
        columns.append('Main Image Links')
    additional_numbers = selected_additional_image_numbers(selected_field_keys)
    for number in additional_numbers:
        columns.append(f'Additional Image {number}')
    if expand_extra_images and additional_numbers:
        start = max(max(additional_numbers), 5) + 1
        for number in range(start, max_additional_image_count + 1):
            columns.append(f'Additional Image {number}')
    return columns


def build_output_row(
    sku: str,
    display_url: str,
    bundle: ExtractionBundle,
    selected_field_keys: List[str],
    max_bullet_count: int,
    max_additional_image_count: int,
    expand_extra_bullets: bool,
    expand_extra_images: bool,
) -> Dict[str, str]:
    row: Dict[str, str] = {'SKU': sku, 'URL': display_url}

    if 'title' in selected_field_keys:
        row['Title'] = bundle.title or NOT_FOUND_TEXT
    if 'description' in selected_field_keys:
        row['Description'] = bundle.description or NOT_FOUND_TEXT

    bullets = bundle.bullets or []
    bullet_numbers = selected_bullet_numbers(selected_field_keys)
    for number in bullet_numbers:
        row[f'Bullet {number}'] = bullets[number - 1] if len(bullets) >= number else NOT_FOUND_TEXT
    if expand_extra_bullets and bullet_numbers:
        start = max(max(bullet_numbers), 5) + 1
        for number in range(start, max_bullet_count + 1):
            row[f'Bullet {number}'] = bullets[number - 1] if len(bullets) >= number else NOT_FOUND_TEXT

    for key in [
        'gender',
        'count_per_pack',
        'multipack',
        'total_count',
        'size',
        'serving_size',
        'color',
        'flavour',
        'product_form',
        'ingredient_statement',
        'dosage',
        'directions',
        'instructions',
        'stop_use_indications',
        'health_concern',
    ]:
        if key in selected_field_keys:
            row[FIELD_LABELS[key]] = bundle.field_values.get(key, '') or NOT_FOUND_TEXT

    images = bundle.images or []
    if 'main_image_links' in selected_field_keys:
        row['Main Image Links'] = images[0] if len(images) >= 1 else NOT_FOUND_TEXT
    additional_numbers = selected_additional_image_numbers(selected_field_keys)
    for number in additional_numbers:
        row[f'Additional Image {number}'] = images[number] if len(images) >= number + 1 else NOT_FOUND_TEXT
    if expand_extra_images and additional_numbers:
        start = max(max(additional_numbers), 5) + 1
        for number in range(start, max_additional_image_count + 1):
            row[f'Additional Image {number}'] = images[number] if len(images) >= number + 1 else NOT_FOUND_TEXT

    return row


def build_results_dataframe(
    scraped_rows: List[Tuple[str, str, ExtractionBundle]],
    selected_field_keys: List[str],
    expand_extra_bullets: bool,
    expand_extra_images: bool,
) -> pd.DataFrame:
    max_bullet_count = 0
    max_additional_image_count = 0
    if expand_extra_bullets:
        max_bullet_count = max((len(bundle.bullets or []) for _, _, bundle in scraped_rows), default=0)
    if expand_extra_images:
        max_additional_image_count = max((max(len(bundle.images or []) - 1, 0) for _, _, bundle in scraped_rows), default=0)

    columns = build_output_columns(
        selected_field_keys,
        max_bullet_count=max_bullet_count,
        max_additional_image_count=max_additional_image_count,
        expand_extra_bullets=expand_extra_bullets,
        expand_extra_images=expand_extra_images,
    )
    if not scraped_rows:
        return pd.DataFrame(columns=columns)

    output_rows: List[Dict[str, str]] = []
    for sku, display_url, bundle in scraped_rows:
        row = build_output_row(
            sku=sku,
            display_url=display_url,
            bundle=bundle,
            selected_field_keys=selected_field_keys,
            max_bullet_count=max_bullet_count,
            max_additional_image_count=max_additional_image_count,
            expand_extra_bullets=expand_extra_bullets,
            expand_extra_images=expand_extra_images,
        )
        output_rows.append(row)

    results_df = pd.DataFrame(output_rows)
    for column in columns:
        if column not in results_df.columns:
            results_df[column] = NOT_FOUND_TEXT if column not in {'SKU', 'URL'} else ''
    return results_df[columns].fillna(NOT_FOUND_TEXT)


def build_output_bytes(results_df: pd.DataFrame, failures: List[Dict[str, str]]) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        results_df.to_excel(writer, sheet_name='Extracted Data', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Extracted Data']

        header_fill = PatternFill(fill_type='solid', fgColor='DCEBFF')
        header_font = Font(bold=True, color='0F172A')
        wrap_alignment = Alignment(vertical='top', wrap_text=True)
        missing_fill = PatternFill(fill_type='solid', fgColor='FDE9EF')

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical='center')

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap_alignment
                if clean_text(cell.value) == NOT_FOUND_TEXT:
                    cell.fill = missing_fill

        width_defaults = {
            'SKU': 18,
            'URL': 42,
            'Title': 42,
            'Description': 72,
            'Ingredient Statement': 66,
            'Directions': 66,
            'Instructions': 66,
            'Stop Use Indications': 66,
            'Dosage': 56,
            'Main Image Links': 54,
        }
        for col_idx, col_name in enumerate(results_df.columns, start=1):
            if col_name in width_defaults:
                width = width_defaults[col_name]
            elif str(col_name).startswith('Bullet '):
                width = 52
            elif str(col_name).startswith('Additional Image '):
                width = 54
            else:
                width = min(max(len(str(col_name)) + 4, 18), 46)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = width

        for row_idx in range(2, worksheet.max_row + 1):
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                value = clean_text(cell.value)
                if value.startswith('http://') or value.startswith('https://'):
                    cell.hyperlink = value
                    cell.style = 'Hyperlink'
                    if value == NOT_FOUND_TEXT:
                        cell.fill = missing_fill

        worksheet.freeze_panes = 'A2'
        worksheet.auto_filter.ref = worksheet.dimensions

        if failures:
            failures_df = pd.DataFrame(failures)
            failures_df.to_excel(writer, sheet_name='Run Log', index=False)
            log_sheet = writer.sheets['Run Log']
            for cell in log_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            for col_idx, col_name in enumerate(failures_df.columns, start=1):
                width = min(max(len(col_name) + 4, 18), 60)
                log_sheet.column_dimensions[get_column_letter(col_idx)].width = width
            for row in log_sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = wrap_alignment

        workbook.properties.creator = APP_TITLE
    output.seek(0)
    return output.getvalue()


def build_template_bytes(rows: int = DEFAULT_ROWS) -> bytes:
    template_df = build_blank_input_df(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        template_df.to_excel(writer, sheet_name='Input Template', index=False)
        worksheet = writer.sheets['Input Template']
        worksheet.freeze_panes = 'A2'
        worksheet.column_dimensions['A'].width = 22
        worksheet.column_dimensions['B'].width = 46
        for cell in worksheet[1]:
            cell.fill = PatternFill(fill_type='solid', fgColor='E7F1FF')
            cell.font = Font(bold=True)
    output.seek(0)
    return output.getvalue()


def kpi_card(label: str, value: str, sub: str = '') -> None:
    st.markdown(
        f"""
        <div class="mini-kpi">
            <div class="label">{label}</div>
            <div class="value">{value}</div>
            <div class="sub">{sub}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def reset_table_only() -> None:
    st.session_state.input_df = build_blank_input_df(DEFAULT_ROWS)
    st.session_state.results_df = pd.DataFrame(columns=default_output_columns())
    st.session_state.last_failures = []
    st.session_state.uploaded_signature = None
    st.session_state.row_count = DEFAULT_ROWS
    st.rerun()


def logout_user() -> None:
    st.session_state.authenticated = False
    st.session_state.user_email = ''
    st.session_state.user_name = ''
    st.session_state.input_df = build_blank_input_df(DEFAULT_ROWS)
    st.session_state.results_df = pd.DataFrame(columns=default_output_columns())
    st.session_state.last_failures = []
    st.session_state.uploaded_signature = None
    st.session_state.file_name = DEFAULT_OUTPUT_FILENAME
    st.session_state.row_count = DEFAULT_ROWS
    st.rerun()


def render_login_page() -> None:
    inject_login_css()
    left, center, right = st.columns([1.3, 1.0, 1.3], gap='large')
    with center:
        st.markdown('<div class="login-card-frame">', unsafe_allow_html=True)
        with st.container(border=False):
            st.markdown(f'<div class="login-title">{APP_TITLE}</div>', unsafe_allow_html=True)
            st.markdown(
                '<div class="login-subtitle">Only approved @pattern.com users can access this tool</div>',
                unsafe_allow_html=True,
            )
            with st.form('login_form', clear_on_submit=False):
                email = st.text_input('Email Address', placeholder='you@pattern.com', key='login_email')
                password = st.text_input('Password', type='password', placeholder='Enter password', key='login_password')
                submitted = st.form_submit_button('Sign In', type='primary', use_container_width=True)
            if submitted:
                ok, message = authenticate_user(email, password)
                if ok:
                    normalized_email = normalize_email(email)
                    st.session_state.authenticated = True
                    st.session_state.user_email = normalized_email
                    st.session_state.user_name = get_display_name(normalized_email)
                    st.rerun()
                else:
                    st.error(message)
            st.markdown(
                '<div class="login-note">New to this page? Please contact Pratik Adsare for creating your login credential</div>',
                unsafe_allow_html=True,
            )
        st.markdown('</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="login-footer">© Designed and Developed by Pratik Adsare</div>',
        unsafe_allow_html=True,
    )

def render_header(user_name: str, user_email: str) -> None:
    left, right = st.columns([4.1, 1.2], gap='large')
    with left:
        st.markdown(
            f"""
            <div class="hero-card">
                <div class="hero-title">{APP_TITLE}</div>
                <div class="hero-subtitle">{APP_SUBTITLE}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with right:
        st.markdown(
            f"""
            <div class="profile-card">
                <div class="profile-label">Signed in</div>
                <div class="profile-name">Welcome {clean_text(user_name) or 'User'}</div>
                <div class="profile-email">{clean_text(user_email)}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.button('Log Out', key='logout_button'):
            logout_user()


def render_attribute_selector() -> Tuple[List[str], bool, bool]:
    with st.container(border=True):
        st.markdown('<div class="attributes-panel-anchor"></div>', unsafe_allow_html=True)
        st.markdown('### Required attributes')
        st.caption('Select only the PDP fields you want in the export. SKU and URL are always included.')
        for group_name, items in FIELD_GROUPS:
            st.markdown(f'#### {group_name}')
            columns = st.columns(ATTRIBUTES_COLUMNS_PER_GROUP)
            for idx, (field_key, label, _default) in enumerate(items):
                with columns[idx % ATTRIBUTES_COLUMNS_PER_GROUP]:
                    st.checkbox(label, key=f'sel_{field_key}')
        st.markdown('#### Dynamic options')
        dynamic_cols = st.columns(2)
        with dynamic_cols[0]:
            st.checkbox('Expand extra bullets', key='expand_extra_bullets')
        with dynamic_cols[1]:
            st.checkbox('Expand extra images', key='expand_extra_images')
        st.markdown(
            """
            <div class="soft-note">
                • Only selected fields are added to the Excel output.<br>
                • If a selected field is unavailable on the PDP, the file writes <strong>Unable to find on PDP</strong> and highlights that cell in light pink.<br>
                • Extra bullets and extra images are appended only when their separate expand option is enabled.<br>
                • Failed URLs are still included in the export and also recorded on the <strong>Run Log</strong> sheet.
            </div>
            """,
            unsafe_allow_html=True,
        )
    return (
        selected_field_keys_from_state(),
        bool(st.session_state.expand_extra_bullets),
        bool(st.session_state.expand_extra_images),
    )


def run_scrape(
    input_rows: pd.DataFrame,
    selected_field_keys: List[str],
    expand_extra_bullets: bool,
    expand_extra_images: bool,
) -> Tuple[pd.DataFrame, List[Dict[str, str]]]:
    total = len(input_rows)
    progress_box = st.container()
    scraped_rows: List[Tuple[str, str, ExtractionBundle]] = []
    failures: List[Dict[str, str]] = []
    status_placeholder = progress_box.empty()
    progress_bar = progress_box.progress(0.0, text='Preparing deep scraper...')
    live_log = progress_box.empty()
    live_lines: List[str] = []

    for idx, row in enumerate(input_rows.to_dict(orient='records'), start=1):
        sku = clean_text(row.get('SKU', ''))
        display_url = clean_text(row.get('Walmart URL', ''))
        url = normalize_url(display_url)

        status_placeholder.markdown(
            f"""
            <div class="status-box">
                <div class="status-line"><strong>Processing {idx} of {total}</strong></div>
                <div class="status-line">Mode: Deep scan</div>
                <div class="status-line">SKU: <code>{sku or '—'}</code></div>
                <div class="status-line">URL: {display_url or '—'}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        progress_bar.progress(idx / total, text=f'Processing {idx} of {total}')

        bundle = scrape_listing(url, selected_field_keys)
        scraped_rows.append((sku, display_url, bundle))
        if bundle.error:
            failures.append({'SKU': sku, 'URL': display_url, 'Error': bundle.error})
            live_lines.append(f"{idx}. {sku or '[no SKU]'} - warning: {bundle.error}")
        else:
            completeness = compute_bundle_completeness(bundle, selected_field_keys)
            live_lines.append(f"{idx}. {sku or '[no SKU]'} - extracted successfully ({completeness:.0%} coverage)")
        live_log.code('\n'.join(live_lines[-10:]), language='text')
        if idx < total:
            time.sleep(REQUEST_DELAY_SECONDS)

    progress_bar.progress(1.0, text='Complete')
    success_count = total - len(failures)
    status_placeholder.success(f'Complete - extracted {success_count} of {total} listing(s).')
    results_df = build_results_dataframe(
        scraped_rows,
        selected_field_keys=selected_field_keys,
        expand_extra_bullets=expand_extra_bullets,
        expand_extra_images=expand_extra_images,
    )
    return results_df, failures

def render_dashboard() -> None:
    render_header(st.session_state.user_name, st.session_state.user_email)

    top_left, top_mid, top_right = st.columns([0.95, 1.1, 1.5])
    with top_right:
        uploaded_file = st.file_uploader(
            'Optional: upload Excel/CSV with SKU and URL columns',
            type=['xlsx', 'xls', 'csv'],
            help='If you upload a file, the editor below will be filled automatically.',
        )
    if uploaded_file is not None:
        signature = (uploaded_file.name, uploaded_file.size)
        if signature != st.session_state.uploaded_signature:
            try:
                uploaded_df, upload_message = parse_uploaded_dataframe(uploaded_file)
                desired_rows = max(DEFAULT_ROWS, len(uploaded_df))
                st.session_state.row_count = desired_rows
                st.session_state.input_df = ensure_row_count(uploaded_df, desired_rows)
                st.session_state.uploaded_signature = signature
                st.success(upload_message)
            except Exception as exc:
                st.error(f'Could not read the uploaded file: {exc}')

    with top_left:
        st.number_input(
            'Rows to show',
            min_value=1,
            max_value=MAX_ROWS,
            step=1,
            key='row_count',
            help='Default is 10. You can scale this up to 1000 rows.',
        )
    with top_mid:
        file_name = st.text_input(
            'Output file name',
            value=st.session_state.file_name,
            help='This name will be used for the downloaded Excel file.',
        )
        st.session_state.file_name = file_name

    st.session_state.input_df = ensure_row_count(st.session_state.input_df, int(st.session_state.row_count))
    editor_df = coerce_input_df(st.session_state.input_df)

    ready_rows = int(((editor_df['SKU'].map(bool)) & (editor_df['Walmart URL'].map(bool))).sum())
    completed_rows = len(st.session_state.results_df)
    failed_rows = len(st.session_state.last_failures)

    kpi1, kpi2, kpi3, kpi4 = st.columns(4)
    with kpi1:
        kpi_card('Configured rows', str(len(editor_df)), f'Default {DEFAULT_ROWS} · max {MAX_ROWS}')
    with kpi2:
        kpi_card('Ready to scrape', str(ready_rows), 'Rows with both SKU and Walmart URL')
    with kpi3:
        kpi_card('Last run output', str(completed_rows), 'Rows exported in the last scrape')
    with kpi4:
        kpi_card('Last run failures', str(failed_rows), 'See Run Log sheet if any rows fail')

    if len(editor_df) <= SCROLL_AFTER_ROWS:
        visible_rows = len(editor_df)
    elif len(editor_df) <= MAX_VISIBLE_ROWS:
        visible_rows = SCROLL_AFTER_ROWS
    else:
        visible_rows = MAX_VISIBLE_ROWS
    editor_height = 46 + max(1, visible_rows) * TABLE_ROW_HEIGHT
    dashboard_panel_height = max(DASHBOARD_PANEL_MIN_HEIGHT, editor_height + 190)
    inject_panel_height_css(dashboard_panel_height)

    selected_field_keys: List[str] = []
    expand_extra_bullets = False
    expand_extra_images = False
    start_clicked = False

    left, right = st.columns([2.15, 1.15], gap='large')
    with left:
        with st.container(border=True):
            st.markdown('<div class="input-panel-anchor"></div>', unsafe_allow_html=True)
            st.markdown('### Input grid')
            st.caption('Paste directly from Excel or type manually. The table stays compact and becomes scrollable for larger batches.')
            edited_df = st.data_editor(
                editor_df,
                hide_index=True,
                height=editor_height,
                use_container_width=True,
                num_rows='fixed',
                column_config={
                    'SKU': st.column_config.TextColumn('SKU', width='medium', help='Your internal SKU or item identifier.'),
                    'Walmart URL': st.column_config.TextColumn('Walmart URL', width='large', help='Paste a Walmart product page URL.'),
                },
                key='input_editor',
            )
            st.session_state.input_df = coerce_input_df(edited_df)
            if len(editor_df) > SCROLL_AFTER_ROWS:
                st.markdown(
                    f'<div class="table-note">Showing a compact scroll area for {len(editor_df)} rows. Scroll inside the grid to review everything.</div>',
                    unsafe_allow_html=True,
                )
            action_col1, action_col2, action_col3 = st.columns([1.2, 0.95, 1.3])
            with action_col1:
                start_clicked = st.button('Start Scraping', type='primary', use_container_width=True)
            with action_col2:
                if st.button('Reset Table', use_container_width=True):
                    reset_table_only()
            with action_col3:
                template_name = f'walmart_input_template_{DEFAULT_ROWS}_rows.xlsx'
                st.download_button(
                    label='Download Blank Input Template',
                    data=build_template_bytes(DEFAULT_ROWS),
                    file_name=template_name,
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    use_container_width=True,
                )

    with right:
        selected_field_keys, expand_extra_bullets, expand_extra_images = render_attribute_selector()

    if start_clicked:
        current_input = coerce_input_df(st.session_state.input_df)
        current_input['Walmart URL'] = current_input['Walmart URL'].map(normalize_url)
        ready_mask = current_input['SKU'].map(bool) & current_input['Walmart URL'].map(bool)
        valid_rows = current_input.loc[ready_mask].copy()

        if not selected_field_keys:
            st.warning('Select at least one attribute before starting the scrape.')
        else:
            invalid_url_mask = ~valid_rows['Walmart URL'].map(looks_like_walmart_url)
            if invalid_url_mask.any():
                bad_urls = valid_rows.loc[invalid_url_mask, 'Walmart URL'].tolist()[:3]
                st.error('Only Walmart product URLs are supported. Please fix these example URL(s): ' + ', '.join(bad_urls))
            elif valid_rows.empty:
                st.warning('Add at least one row with both SKU and Walmart URL before starting.')
            else:
                incomplete_count = len(current_input[(current_input['SKU'].map(bool) ^ current_input['Walmart URL'].map(bool))])
                if incomplete_count:
                    st.info(f'Ignoring {incomplete_count} incomplete row(s) that do not have both SKU and URL.')
                results_df, failures = run_scrape(
                    valid_rows,
                    selected_field_keys=selected_field_keys,
                    expand_extra_bullets=expand_extra_bullets,
                    expand_extra_images=expand_extra_images,
                )
                st.session_state.results_df = results_df
                st.session_state.last_failures = failures

    if not st.session_state.results_df.empty:
        st.markdown('### Results')
        result_cols = st.columns([1.2, 1.2, 1.1])
        with result_cols[0]:
            st.success(f'Rows ready for export: {len(st.session_state.results_df)}')
        with result_cols[1]:
            st.info(f'Failures recorded: {len(st.session_state.last_failures)}')
        with result_cols[2]:
            file_stub = slugify_filename(st.session_state.file_name)
            st.download_button(
                label='Download Excel Output',
                data=build_output_bytes(st.session_state.results_df, st.session_state.last_failures),
                file_name=f'{file_stub}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                type='primary',
                use_container_width=True,
            )
        preview_tabs = st.tabs(['Output Preview', 'Failures'])
        with preview_tabs[0]:
            st.dataframe(
                st.session_state.results_df,
                use_container_width=True,
                hide_index=True,
                height=min(560, 46 + min(len(st.session_state.results_df), 10) * 40),
            )
        with preview_tabs[1]:
            if st.session_state.last_failures:
                st.dataframe(pd.DataFrame(st.session_state.last_failures), use_container_width=True, hide_index=True)
            else:
                st.success('No failures in the last run.')


def main() -> None:
    inject_css()
    init_state()
    if not st.session_state.authenticated:
        render_login_page()
        return
    render_dashboard()


if __name__ == '__main__':
    main()
